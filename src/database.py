from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import collections.abc
import json
import pickle
import pathlib
import inspect
import copy


class TransPair(collections.UserList):
    """A translation pair.

    A pair of string elements, to store sentence translation.
    Empty strings are invalid.
    """
    def __init__(self, item0, item1):
        if not isinstance(item0, str) and not isinstance(item1, str):
            raise TypeError("string element expected.")
        elif (item0 and item1) == "":
            raise ValueError("cannot store empty strings.")
        super(TransPair, self).__init__([item0, item1])

    def __getattribute__(self, attr):
        """Prevent 'private' attribute access"""
        private_super = [
            "append", "extend", "insert",
            "pop", "remove", "clear", "sort"
        ]
        if attr in private_super:
            raise AttributeError
        else:
            return super(TransPair, self).__getattribute__(attr)

    def __setitem__(self, key, value):
        if not isinstance(value, str):
            raise TypeError("string element expected.")
        elif len(value) == 0:
            raise ValueError("cannot store empty strings.")
        super(TransPair, self).__setitem__(key, value)

    def switch(self):
        """Switch the elements. Invert the TransPair order."""
        super(TransPair, self).reverse()


class TransList(collections.UserList):
    """A list/set of TransPair."""

    def __init__(self, *args):
        """
        :param args: TransPair or lists to construct TransPairs
        """
        super(TransList, self).__init__()
        for arg in args:
            self.append(arg)

    def __str__(self):
        indent = "    "
        trans_list_str = str()
        for i, item in enumerate(super(TransList, self).__iter__()):
            trans_list_str = trans_list_str + f"{indent} {str(item)}"
            if i < super(TransList, self).__len__() - 1:
                trans_list_str = trans_list_str + ", \n"
        return f"{{\n{trans_list_str}\n}}"

    def append(self, tp_key):
        """Adds a new TransPair element into the end of the TransList."""
        if not isinstance(tp_key, TransPair):
            raise TypeError("only TransPair object is accepted.")
        elif super(TransList, self).__contains__(tp_key):
            raise KeyError("TransPair {} already assigned.".format(tp_key))
        else:
            super(TransList, self).append(tp_key)

    def insert(self, i, tp_key):
        # TODO: test this method
        if type(tp_key) != TransPair:
            raise KeyError("the tp_key must be a TransPair object.")
        elif super(TransList, self).__contains__(tp_key):
            raise KeyError("TransPair {} already assigned.".format(tp_key))
        super(TransList, self).insert(i, tp_key)

    def extend(self, other):
        # TODO: test this method
        for trans_pair in other:
            self.append(trans_pair)

    def remove(self, tp_key):
        """Remove a TransPair element by element key."""
        super(TransList, self).remove(tp_key)

    def pop(self, index=-1):
        """Remove the item at the given position in the list, and return it.

        If no index is specified, a.pop() removes and returns the last item in
        the TransList.
        """
        super(TransList, self).pop(index)

    def get_translation(self, sentence, hint=0):
        """Return a the translation.

        Hint is where the search will take place. If 0 will search the sentence on
        all TransPairs elements. if 1 will do the search only on the first TransPairs
        index. If 2 will do the search only on the second TransPairs index.
        :param sentence:
        :param hint:
        :return: str
        """
        if hint == 0 or hint == 1:
            for transpair in self:
                if transpair[0] == sentence:
                    return transpair[1]
        if hint == 0 or hint == 2:
            for transpair in self:
                if transpair[1] == sentence:
                    return transpair[0]


class TransDatabase(dict):
    """Creates, load and manipulate TransLists.

    The TransDatabase are a dictionary that store and manipulate named TransList
    that manipulate TransPairs objects, and are constructed with some helpful
    attributes to organize your TransLists, load and manipulate TransPairs.
    In a TransDatabase, the TransList has a unique key to uniquely identify
    the TransList, but his content can be the same of other TransList.
    When you initialize a empty TransDatabase you need to identify a
    first_lang("e.g. EN-US") and second_lang("e.g. DE-CH") attribute. These are
    mere index for you to know the order in which you have to store the
    sentences, but not that does not stop the user store them wrong.
    """
    def __init__(self, first_lang, second_lang, **kwargs):
        super().__init__()
        self.info = {"language": TransPair(first_lang, second_lang),
                     "creator": "Daniel Daninsky",
                     "script": "duobreaker v0.1"}
        for trans_list_name, trans_pairs in kwargs.items():
            self.add(trans_list_name, trans_pairs)

    def __setitem__(self, trans_list_name, trans_list):
        """Set self[key] to value.

        Caution: note that __setitem__ can overwrite data. So if you want add
        a new key with a overwrite garanty use the add method.
        """
        if not isinstance(trans_list_name, str):
            raise TypeError("invalid type")
        elif not isinstance(trans_list, TransList):
            TypeError("object isn't a TransList")
        else:
            super(TransDatabase, self).__setitem__(trans_list_name, trans_list)

    def add(self, trans_list_name, trans_list=None):
        """Add TransList with its name in the end of TransDatabase.

        If TransList is empty the effect are the same of the method add_trans_list.
        Add are recomended to add new TransList. Add has overwrite guard that
        __setitem__ doesn't.
        """
        if trans_list_name in self:
            raise KeyError("TransList already exists")
        if trans_list == None:
            trans_list = TransList()
        else:
            self.__setitem__(trans_list_name, trans_list)

    def get_translation(self, key, sentence, lang=None):
        """Get the translation."""
        if lang == self.langs[0]:
            return self[key].get_translation(sentence, 1)
        elif lang == self.langs[1]:
            return self[key].get_translation(sentence, 2)
        else:
            return self[key].get_translation(sentence, 0)

    def getbyindex(self, index):
        """Return TransList by index."""
        if index < 0 or index > len(self):
            raise IndexError("TransDatabase index out of range")
        for i, item in enumerate(self):
            if i == index:
                return self[item]

    def save(self, file, overwrite=False, extension="xlsx"):
        if not isinstance(extension, str):
            raise TypeError("invalid type")
        extension = extension.casefold()
        if extension == "xlsx":
            self.__xlsx_save(file)
        elif extension == "json":
            self.__json_save(file)
        else:
            raise ValueError("invalid extension type. xlsx or json")

    @classmethod
    def fromfile(cls, file):
        """Load a from disk a .xlsx translation database.

        :param file: Path-like object where the database will be opened
        :return: TransDatabase
        """
        if not isinstance(file, str):
            raise TypeError("invalid type")
        extension = file[-4:]
        extension = extension.casefold()    # <- this is fucking bullshit
        if extension == "xlsx":
            # TODO: add a attribute support to the files
            kwargs = cls.__xlsx_load(file)
            first_lang = "EN-US"   # TODO: hardcoded temporariamente
            second_lang = "PT-BR"
            return cls(first_lang, second_lang, **kwargs)
        elif extension == "json":
            cls.__json_load(file)
        else:
            raise ValueError("invalid extension file. xlsx or json")

    def change_lang_attrs(self, first_lang, second_lang):
        """Change the language attributes."""
        self.info["language"] = TransPair(first_lang, second_lang)

    def __xlsx_save(self, file):
        # TODO: improve style and add a info style
        workbook = Workbook()
        worksheet = workbook.active  # current working spreadsheet
        workbook.remove(worksheet)
        def sheet_decor(worksheet):
            """Sheet style fo here, font, size, etc."""
            default_font = Font(name='Arial', size=12)
            header_font = Font(name='Arial', size=14, bold=True)
            worksheet.column_dimensions['A'].font = default_font
            worksheet.column_dimensions['B'].font = default_font
            worksheet['A1'].font = header_font
            worksheet['B1'].font = header_font
            cell_width = 60
            cell_height = 20
            worksheet.column_dimensions['A'].width = cell_width
            worksheet.column_dimensions['B'].width = cell_width
            worksheet.row_dimensions[1].height = cell_height
        # add content
        for key, value in self.items():
            worksheet = workbook.create_sheet(key)
            sheet_decor(worksheet)
            for num, trans_pair in enumerate(value, start=1):
                worksheet.cell(row=num, column=1).value = trans_pair[0]
                worksheet.cell(row=num, column=2).value = trans_pair[1]
        # add info
        worksheet_info = workbook.create_sheet("info")
        for num, (key, value) in enumerate(self.info.items(), start=1):
            worksheet_info.cell(row=num, column=1).value = key
            worksheet_info.cell(row=num, column=2).value = repr(value)   # repr to serialize the object
        workbook.save(file)

    def __json_save(self, file):
        # TODO: implement
        pass

    @staticmethod
    def __xlsx_load(file):
        workbook = load_workbook(file)
        kwargs = {}
        for spreadsheet in workbook.worksheets:
            trans_list = TransList()
            spreadsheet_name = spreadsheet.title
            for row in spreadsheet.iter_rows(min_row=1, max_col=2, values_only=True):
                # TODO: add ordered sequence suport to the TransPair constructor
                trans_list.append(TransPair(row[0], row[1]))
            kwargs[spreadsheet_name] = trans_list
        return kwargs

    def __json_load(self, file):
        # TODO: implement
        pass


def test_db2():
    my_translist = TransList(("hi", "oi"))
    my_translist2 = TransList(["car", "carro"])
    new_translist = TransList()
    my_trans_pair = TransPair("olá", "hello")
    new_translist.extend(my_translist)
    new_translist.extend(my_translist2)

    print(my_translist, my_translist2, sep="\n")


def tl_contains_test():
    hi = TransPair("hi", "oi")
    my_trans_list = TransList(hi)
    my_trans_list.append(hi)

def td_test():
    my_tl = TransList()
    translations = [["car", "carro"], ["water", "água"]]
    for phrases in translations:
        my_tl.append(TransPair(phrases[0], phrases[1]))

    my_database = TransDatabase("en", "pt")
    my_database.add("saudações", my_tl)
    my_database.add("phrases")
    my_database["daniel"] = my_tl

    my_tl_back = my_database["saudações"]

    print(len(my_tl))
    print(len(my_database["daniel"]))
    print(type(my_database["daniel"]))
    print(my_tl)
    print(my_database)


if __name__ == '__main__':
    my_tdb = TransDatabase.fromfile(r"../database/Acidentes/Acidentes_en_to_pt_dictionary.xlsx")
    my_tdb.save("test.xlsx")
    print(my_tdb)
